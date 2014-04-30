from django.http import HttpResponse, HttpResponseGone
from django.template.defaultfilters import slugify
import os
import mimetypes
import tempfile
from datetime import date
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Color, Fill


def send_file(path, filename=None, mimetype=None):

    if filename is None:
        filename = os.path.basename(path)

    if mimetype is None:
        mimetype, encoding = mimetypes.guess_type(filename)

    response = HttpResponse(mimetype=mimetype)
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    response.write(file(path, "rb").read())
    return response


def export_table(table, format='xlsx'):
    """ Get a django_tables2 table and generate a spreadsheet
    with its data and send the file to a response
    """

    if format != 'xlsx':
        return HttpResponseGone("Format not implemented")

    wb = Workbook()
    ws = wb.active
    ws.title = table.verbose_name
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1

    # header style
    header_fill = Fill()
    header_fill.start_color.index = Color.BLUE
    header_fill.end_color.index = Color.BLUE
    header_fill.fill_type = Fill.FILL_SOLID

    column_widths = []
    col = 0
    for column in table.columns:
        c = ws.cell(row=0, column=col)
        c.value = column.header.encode('utf8')
        c.style.fill = header_fill
        c.style.font.color.index = Color.WHITE
        c.style.font.bold = True
        c.style.font.size = 12
        column_widths.append(len(unicode(c.value)))
        col = col + 1

    row = 1
    for obj in table.rows:
        col = 0
        for value in obj:
            c = ws.cell(row=row, column=col)
            c.value = unicode(value)
            if len(unicode(c.value)) > column_widths[col]:
                column_widths[col] = len(c.value)
            col = col + 1
        row = row + 1

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 2

    suffix = '.xlsx'
    tf = tempfile.NamedTemporaryFile(suffix=suffix)
    wb.save(tf.name)

    filename = "{date:%Y-%m-%d}-{name}{suffix}".format(
        date=date.today(),
        name=slugify(ws.title),
        suffix=suffix)

    return send_file(
        path=tf.name,
        filename=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
